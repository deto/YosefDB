import openpyxl;
import datetime;
from .models import Sample, Upload, UnvalidatedSample, UnvalidatedUpload;


def parseBool(in_string):
    """
    Reads in_string and parses it to a boolean value
    :param in_string: string to be parsed
    :return: True or False
    """
    in_string = in_string.lower();
    if(in_string.find("yes") > -1 or in_string.find("true") > 1 or in_string=="y" or in_string=="1"):
        return True;
    elif(in_string.find("no") > -1 or in_string.find("false") > 1 or in_string=="n" or in_string=="0"):
        return False;
    else:
        raise ValueError("Can't Parse Boolean: " + in_string);

def parseDate(in_string):
    """
    Reads in_string and outputs Date object
    Expects format <month>/<day>/<year>
    :param in_string: string to be read
    :return: datetime.Date object
    """

    if(type(in_string) is str):
        mdy = [int(x) for x in in_string.split('/')];
        outDate = datetime.date(mdy[2], mdy[0], mdy[1]);
    if(type(in_string) is datetime.datetime):
        outDate = in_string.date();

    return outDate;



def handle_uploaded_file(uploadedBy, file):
    wb = openpyxl.load_workbook(file);
    ws = wb.active;

    thisUpload = UnvalidatedUpload();
    thisUpload.UploadedBy = uploadedBy;

    #Read in the column headers.  Create a mapping to Sample parameters
    col_map = dict();
    unrecognized_columns = list();
    for i in xrange(ws.get_highest_column()):
        headerText = ws.cell(row=0, column=i).value;
        headerText = headerText.lower().replace(" ","");
        headerText = headerText.replace("_","");
        if(headerText == "organism"):
            col_map.update({"organism": i});
        elif(headerText == "tissue"):
            col_map.update({"tissue": i});
        elif(headerText == "experimentname"):
            col_map.update({"experimentname": i});
        elif(headerText == "datatype"):
            col_map.update({"datatype": i});
        elif(headerText == "genomeversion"):
            col_map.update({"genomeversion": i});
        elif(headerText == "experimentdescription"):
            col_map.update({"experimentdescription": i});
        elif(headerText == "singlecell"):
            col_map.update({"singlecell": i});
        elif(headerText == "publisheddata"):
            col_map.update({"publisheddata": i});
        elif(headerText == "date"):
            col_map.update({"date": i});
        elif(headerText == "batch"):
            col_map.update({"batch": i});
        elif(headerText == "sequencingtech"):
            col_map.update({"sequencingtech": i});
        elif(headerText == "rawfileloc"):
            col_map.update({"rawfileloc": i});
        elif(headerText == "specificrawfileloc"):
            col_map.update({"specificrawfileloc": i});
        elif(headerText == "localbackuploc"):
            col_map.update({"localbackuploc": i});
        elif(headerText == "remotebackuploc"):
            col_map.update({"remotebackuploc": i});
        elif(headerText == "processedfileloc"):
            col_map.update({"processedfileloc": i});
        elif(headerText == "conditioncode"):
            col_map.update({"conditioncode": i});
        elif(headerText == "timecode"):
            col_map.update({"timecode": i});
        elif(headerText == "batchcode"):
            col_map.update({"batchcode": i});
        elif(headerText == "idcode"):
            col_map.update({"idcode": i});
        elif(headerText == "controlid"):
            col_map.update({"controlid": i});
        elif(headerText == "preprocessdate"):
            col_map.update({"preprocessdate": i});
        elif(headerText == "comments"):
            col_map.update({"comments": i});
        else:
            unrecognized_columns.append(ws.cell(row=0, column=i).value);

    #Iterate through all the other rows and create Samples for each row
    #An empty cell has a .value of None
    samples_to_add = list();
    for i,row in enumerate(ws.rows):
        if(i==0): continue;
        newSample = UnvalidatedSample();
        if(col_map.has_key("organism")):
            cellVal = row[col_map["organism"]].value;
            if(cellVal is not None): newSample.Organism = cellVal;
        if(col_map.has_key("tissue")):
            cellVal = row[col_map["tissue"]].value;
            if(cellVal is not None): newSample.Tissue = cellVal;
        if(col_map.has_key("experimentname")):
            cellVal = row[col_map["experimentname"]].value;
            if(cellVal is not None): newSample.ExperimentName = cellVal;
        if(col_map.has_key("datatype")):
            cellVal = row[col_map["datatype"]].value;
            if(cellVal is not None): newSample.DataType = cellVal;
        if(col_map.has_key("genomeversion")):
            cellVal = row[col_map["genomeversion"]].value;
            if(cellVal is not None): newSample.GenomeVersion = cellVal;
        if(col_map.has_key("experimentdescription")):
            cellVal = row[col_map["experimentdescription"]].value;
            if(cellVal is not None): newSample.ExperimentDescription = cellVal;
        if(col_map.has_key("singlecell")):
            cellVal = row[col_map["singlecell"]].value;
            if(cellVal is not None): newSample.isSingleCell = parseBool(cellVal);
        if(col_map.has_key("publisheddata")):
            cellVal = row[col_map["publisheddata"]].value;
            if(cellVal is not None): newSample.isPublishedData = parseBool(cellVal);
        if(col_map.has_key("date")):
            cellVal = row[col_map["date"]].value;
            if(cellVal is not None): newSample.Date = parseDate(cellVal);
        if(col_map.has_key("batch")):
            cellVal = row[col_map["batch"]].value;
            if(cellVal is not None): newSample.Batch = cellVal;
        if(col_map.has_key("sequencingtech")):
            cellVal = row[col_map["sequencingtech"]].value;
            if(cellVal is not None): newSample.SequencingTech = cellVal;
        if(col_map.has_key("rawfileloc")):
            cellVal = row[col_map["rawfileloc"]].value;
            if(cellVal is not None): newSample.RawFileLoc = cellVal;
        if(col_map.has_key("specificrawfileloc")):
            cellVal = row[col_map["specificrawfileloc"]].value;
            if(cellVal is not None): newSample.SpecificRawFileLoc = cellVal;
        if(col_map.has_key("localbackuploc")):
            cellVal = row[col_map["localbackuploc"]].value;
            if(cellVal is not None): newSample.LocalBackupLoc = cellVal;
        if(col_map.has_key("remotebackuploc")):
            cellVal = row[col_map["remotebackuploc"]].value;
            if(cellVal is not None): newSample.RemoteBackupLoc = cellVal;
        if(col_map.has_key("processedfileloc")):
            cellVal = row[col_map["processedfileloc"]].value;
            if(cellVal is not None): newSample.ProcessedFileLoc = cellVal;
        if(col_map.has_key("conditioncode")):
            cellVal = row[col_map["conditioncode"]].value;
            if(cellVal is not None): newSample.ConditionCode = int(cellVal);
        if(col_map.has_key("timecode")):
            cellVal = row[col_map["timecode"]].value;
            if(cellVal is not None): newSample.TimeCode = int(cellVal);
        if(col_map.has_key("batchcode")):
            cellVal = row[col_map["batchcode"]].value;
            if(cellVal is not None): newSample.BatchCode = int(cellVal);
        if(col_map.has_key("idcode")):
            cellVal = row[col_map["idcode"]].value;
            if(cellVal is not None): newSample.IdCode = int(cellVal);
        if(col_map.has_key("controlid")):
            cellVal = row[col_map["controlid"]].value;
            if(cellVal is not None): newSample.ControlId = int(cellVal);
        if(col_map.has_key("preprocessdate")):
            cellVal = row[col_map["preprocessdate"]].value;
            if(cellVal is not None): newSample.PreprocessDate = parseDate(cellVal);
        if(col_map.has_key("comments")):
            cellVal = row[col_map["comments"]].value;
            if(cellVal is not None): newSample.Comments = cellVal;

        newSample.UploadBatch = thisUpload;
        samples_to_add.append(newSample);


    thisUpload.UnrecognizedCols = '\n'.join(unrecognized_columns);
    thisUpload.save();
    UnvalidatedSample.objects.bulk_create(samples_to_add);

    return thisUpload;

def validate_upload(unvalidatedUpload):
    """
    Copies the unvalidated upload and all of its related samples into the
    validated Samples and Uploads tables

    :param unvalidatedUpload: The upload that is to be made valid
    :return:
    """

    newUpload = unvalidatedUpload.create_valid_upload();

    unvalid_samples = unvalidatedUpload.unvalidatedsample_set.all();
    valid_samples = list();
    for unvalid_sample in unvalid_samples:
        new_sample = unvalid_sample.create_valid_sample();
        new_sample.UploadBatch = newUpload;
        valid_samples.append(new_sample);

    newUpload.save();
    Sample.objects.bulk_create(valid_samples);










